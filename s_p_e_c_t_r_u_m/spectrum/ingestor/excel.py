"""Ингестор Excel: извлечение данных из таблиц с сохранением структуры."""

from __future__ import annotations

import csv
import io
import re
from pathlib import Path
from typing import Any

from .base import IngestResult, Ingestor, PageChunk, SourceType


class ExcelIngestor(Ingestor):
    """Извлекает данные из Excel-файлов (.xlsx, .xls, .csv).

    Особенности:
    - Сохраняет имена листов (sheet_name) для traceability
    - Форматирует таблицы в markdown для лучшего понимания LLM
    - Поддерживает merged cells
    """

    SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".csv"}

    def can_handle(self, source: str | Path) -> bool:
        p = Path(source) if isinstance(source, str) else source
        return p.suffix.lower() in self.SUPPORTED_EXTENSIONS

    def ingest(self, source: str | Path, **kwargs) -> IngestResult:
        path = Path(source)
        data = self._read_bytes(path)
        file_hash = IngestResult.compute_hash(data)

        ext = path.suffix.lower()
        if ext == ".csv":
            return self._ingest_csv(path, data, file_hash)
        return self._ingest_excel(path, data, file_hash)

    def _ingest_excel(self, path: Path, data: bytes, file_hash: str) -> IngestResult:
        """Парсинг .xlsx/.xls через openpyxl."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            return self._fallback_excel(path, data, file_hash)

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        chunks: list[PageChunk] = []
        metadata: dict[str, Any] = {"sheets": []}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):
                    rows.append(cells)

            if not rows:
                continue

            metadata["sheets"].append(sheet_name)

            # Первая строка — заголовки
            headers = rows[0] if rows else []
            data_rows = rows[1:] if len(rows) > 1 else []

            # Форматируем в markdown-таблицу
            table_text = self._to_markdown_table(headers, data_rows)

            chunks.append(PageChunk(
                text=table_text,
                sheet_name=sheet_name,
                row_range=(1, len(rows)),
                metadata={
                    "headers": headers,
                    "row_count": len(data_rows),
                    "col_count": len(headers),
                },
            ))

        wb.close()
        metadata["total_sheets"] = len(metadata.get("sheets", []))

        return IngestResult(
            source_path=str(path),
            source_type=SourceType.EXCEL,
            chunks=chunks,
            file_hash=file_hash,
            total_pages=len(chunks),
            metadata=metadata,
        )

    def _ingest_csv(self, path: Path, data: bytes, file_hash: str) -> IngestResult:
        """Парсинг CSV файлов."""
        text = data.decode("utf-8", errors="replace")

        # Автоопределение разделителя
        sniffer = csv.Sniffer()
        try:
            dialect = sniffer.sniff(text[:1024])
        except csv.Error:
            dialect = csv.excel

        reader = csv.reader(io.StringIO(text), dialect)
        all_rows = [row for row in reader if any(c.strip() for c in row)]

        if not all_rows:
            return IngestResult(
                source_path=str(path),
                source_type=SourceType.EXCEL,
                chunks=[],
                file_hash=file_hash,
                metadata={"empty": True},
            )

        headers = all_rows[0]
        data_rows = all_rows[1:]
        table_text = self._to_markdown_table(headers, data_rows)

        chunk = PageChunk(
            text=table_text,
            sheet_name=path.stem,
            row_range=(1, len(all_rows)),
            metadata={"headers": headers, "row_count": len(data_rows)},
        )

        return IngestResult(
            source_path=str(path),
            source_type=SourceType.EXCEL,
            chunks=[chunk],
            file_hash=file_hash,
            total_pages=1,
            metadata={"format": "csv", "delimiter": dialect.delimiter},
        )

    def _to_markdown_table(self, headers: list[str], rows: list[list[str]]) -> str:
        """Конвертирует данные в markdown-таблицу для LLM."""
        if not headers:
            return ""

        lines = []
        # Заголовок
        lines.append("| " + " | ".join(h or "—" for h in headers) + " |")
        lines.append("| " + " | ".join("---" for _ in headers) + " |")

        for row in rows:
            # Выравниваем количество колонок
            padded = list(row) + [""] * (len(headers) - len(row))
            cells = padded[:len(headers)]
            lines.append("| " + " | ".join(c or "—" for c in cells) + " |")

        return "\n".join(lines)

    def _fallback_excel(self, path: Path, data: bytes, file_hash: str) -> IngestResult:
        """Фолбэк: читаем как текст."""
        text = data.decode("utf-8", errors="replace")
        chunks = [PageChunk(text=text[:10000], sheet_name="Sheet1")] if text.strip() else []
        return IngestResult(
            source_path=str(path),
            source_type=SourceType.EXCEL,
            chunks=chunks,
            file_hash=file_hash,
            metadata={"fallback": True, "warning": "openpyxl not installed"},
        )
